import threading
from openpyxl import Workbook, load_workbook

def convert_txt_to_excel(input_file, output_file):
    # Mở file Excel hiện có
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    with open(input_file, 'r', encoding='utf-8-sig') as file:
        for line in file:
            line = line.strip()
            if line:
                try:
                    chinese, vietnamese = line.split('=')
                    worksheet.append([chinese, vietnamese])
                    print(f"Đã chuyển đổi dòng: {line}")
                except ValueError:
                    print(f"Lỗi: Không thể chuyển đổi dòng: {line}. Đã bỏ qua.")

    # Lưu lại các thay đổi vào file Excel
    workbook.save(output_file)

# Tạo một lock để đồng bộ hóa truy cập vào file Excel
lock = threading.Lock()

# Hàm chạy trong mỗi thread
def convert_thread(input_file, output_file):
    # Tạo một workbook mới với tiêu đề cột
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Tiếng Trung", "Tiếng Việt"])
    workbook.save(output_file)

    convert_txt_to_excel(input_file, output_file)

# Ví dụ sử dụng
threads = []
for i in range(1, 51):  # Lặp qua các số từ 1 đến 5 (có thể thay đổi theo nhu cầu)
    input_file = f"VietPhrase_{i}.txt"  # Tệp văn bản đầu vào
    output_file = f"VietPhrase_{i}.xlsx"  # Tệp Excel đầu ra

    thread = threading.Thread(target=convert_thread, args=(input_file, output_file))
    thread.start()
    threads.append(thread)

# Chờ cho tất cả các thread hoàn thành
for thread in threads:
    thread.join()